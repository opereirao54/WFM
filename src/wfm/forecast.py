"""
Forecast module — generates volume & TMO curves from historical data.
Uses IQR (boxplot) method for outlier removal.
"""
import io, math, calendar, datetime as _dt
from typing import List, Dict, Tuple, Optional
import numpy as np

from .models import CurvaIntraday, DiaMes

# ── 30-min time labels ────────────────────────────────────────────────
HORARIOS = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]

# Feriados brasileiros fixos (data fixa todo ano)
_FERIADOS_FIXOS = {
    (1, 1), (4, 21), (5, 1), (9, 7), (10, 12), (11, 2), (11, 15), (12, 25),
}


def _pascoa(ano: int) -> _dt.date:
    """Calcula a data da Páscoa no ano (algoritmo de Butcher/Meeus)."""
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    L = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * L) // 451
    mes = (h + L - 7 * m + 114) // 31
    dia = ((h + L - 7 * m + 114) % 31) + 1
    return _dt.date(ano, mes, dia)


def _feriados_moveis(ano: int) -> set:
    """Retorna tuplas (mes, dia) dos feriados móveis: Carnaval (seg+ter),
    Sexta-feira Santa, Páscoa, Corpus Christi."""
    p = _pascoa(ano)
    carn_ter = p - _dt.timedelta(days=47)  # terça de carnaval
    carn_seg = p - _dt.timedelta(days=48)  # segunda de carnaval
    sext_sta = p - _dt.timedelta(days=2)   # sexta-feira santa
    corpus   = p + _dt.timedelta(days=60)  # Corpus Christi (quinta)
    return {(d.month, d.day) for d in (carn_seg, carn_ter, sext_sta, p, corpus)}


def _is_feriado(d: _dt.date) -> bool:
    """Verifica se a data é feriado brasileiro (fixo ou móvel do ano)."""
    if (d.month, d.day) in _FERIADOS_FIXOS:
        return True
    return (d.month, d.day) in _feriados_moveis(d.year)


def _tipo_dia(d: _dt.date) -> str:
    """Classify a date as Util, Sabado, or Domingo (holidays → Domingo)."""
    if _is_feriado(d):
        return "Domingo"
    dow = d.weekday()
    if dow == 5:
        return "Sabado"
    if dow == 6:
        return "Domingo"
    return "Util"


# ── IQR outlier removal ──────────────────────────────────────────────
def remove_outliers_iqr(values: np.ndarray, factor: float = 1.5):
    """
    Remove outliers using IQR (boxplot) method.
    Returns (cleaned_values, q1, q3, iqr, lower_bound, upper_bound, n_removed).
    """
    if len(values) < 4:
        return values, 0, 0, 0, 0, 0, 0
    q1 = float(np.percentile(values, 25))
    q3 = float(np.percentile(values, 75))
    iqr = q3 - q1
    lower = q1 - factor * iqr
    upper = q3 + factor * iqr
    mask = (values >= lower) & (values <= upper)
    n_removed = int((~mask).sum())
    return values[mask], q1, q3, iqr, lower, upper, n_removed


# ── Parse historical Excel ───────────────────────────────────────────
def get_forecast_template_xlsx() -> bytes:
    """Generate template Excel for historical data upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historico"
    ws.sheet_view.showGridLines = False

    h_fill = PatternFill("solid", fgColor="1a2030")
    h_font = Font(bold=True, color="4f9cf9", size=11)
    inst_font = Font(color="6c7a96", size=10, italic=True)
    data_font = Font(color="cdd6f4", size=10)
    bd = Side(style="thin", color="1e2535")
    border = Border(left=bd, right=bd, top=bd, bottom=bd)
    center = Alignment(horizontal="center")

    # Instructions
    instructions = [
        "INSTRUÇÕES DE PREENCHIMENTO",
        "",
        "Coluna DATA: data no formato AAAA-MM-DD. Ex: 2025-03-15",
        "Coluna HORARIO: horário de 30min (00:00, 00:30, ..., 23:30)",
        "Coluna VOLUME: número de chamadas nesse intervalo",
        "Coluna TMO: tempo médio de operação em SEGUNDOS",
        "",
        "Inclua dados de pelo menos 1 mês (recomendado 3+).",
        "Cada linha = um intervalo de 30 minutos de um dia.",
        "Feriados serão tratados automaticamente como Domingo.",
    ]
    for r, txt in enumerate(instructions, 1):
        c = ws.cell(row=r, column=6, value=txt)
        c.font = inst_font
        if r == 1:
            c.font = Font(color="f9e2af", size=11, bold=True)

    # Headers
    for j, h in enumerate(["DATA", "HORARIO", "VOLUME", "TMO"], 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = center
        c.border = border

    # Example data (3 days × 48 intervals)
    example_dates = ["2025-03-03", "2025-03-04", "2025-03-05"]
    row = 3
    for dt_str in example_dates:
        for h_str in HORARIOS:
            ws.cell(row=row, column=1, value=dt_str).font = data_font
            ws.cell(row=row, column=2, value=h_str).font = data_font
            ws.cell(row=row, column=3, value="").font = data_font
            ws.cell(row=row, column=4, value="").font = data_font
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = center
                ws.cell(row=row, column=col).border = border
            row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["F"].width = 55
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_historico_xlsx(file_bytes: bytes) -> List[dict]:
    """
    Parse historical data Excel.
    Returns list of dicts: {data, horario, volume, tmo, tipo_dia}
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if "Historico" not in wb.sheetnames:
        # Try first sheet
        ws = wb.active
    else:
        ws = wb["Historico"]

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        try:
            data_str = str(row[0]).strip()[:10]
            d = _dt.date.fromisoformat(data_str)
            horario = str(row[1]).strip()[:5]
            vol = float(str(row[2]).replace(",", ".")) if row[2] is not None else 0.0
            tmo = float(str(row[3]).replace(",", ".")) if row[3] is not None else 0.0
            if vol < 0:
                vol = 0.0
            if tmo < 0:
                tmo = 0.0
            records.append({
                "data": data_str,
                "horario": horario,
                "volume": vol,
                "tmo": tmo,
                "tipo_dia": _tipo_dia(d),
                "mes": f"{d.year}-{d.month:02d}",
            })
        except (ValueError, TypeError, IndexError):
            continue

    return records


# ── Forecast: intraday curve ─────────────────────────────────────────
def forecast_curva_intraday(records: List[dict], tipo_dia: str, window: Optional[Tuple[int, int]] = None):
    """
    Generate intraday curve (48 slots) for a day type.
    window: Optional tuple of (start_slot, end_slot) inclusive, e.g. (16, 40) for 08:00 to 20:00.
    Returns (pesos_pct[48], fatores_tmo[48], stats_per_slot).
    """
    # Group by horario
    by_slot = {h: {"vols": [], "tmos": []} for h in HORARIOS}
    for r in records:
        if r["tipo_dia"] != tipo_dia:
            continue
        h = r["horario"]
        if h in by_slot:
            by_slot[h]["vols"].append(r["volume"])
            if r["volume"] > 0 and r["tmo"] > 0:
                by_slot[h]["tmos"].append(r["tmo"])

    pesos = []
    fatores_tmo = []
    stats = []

    for i, h in enumerate(HORARIOS):
        vols = np.array(by_slot[h]["vols"], dtype=float)
        tmos = np.array(by_slot[h]["tmos"], dtype=float)

        # Se estiver fora da janela de operação permitida, forçamos 0 (antes do IQR)
        if window is not None and not (window[0] <= i < window[1]):
            vols = np.array([], dtype=float)
            tmos = np.array([], dtype=float)

        # Clean volume outliers
        if len(vols) >= 4:
            vol_clean, q1v, q3v, iqrv, lv, uv, nrv = remove_outliers_iqr(vols)
        else:
            vol_clean = vols
            q1v = float(np.percentile(vols, 25)) if len(vols) > 0 else 0
            q3v = float(np.percentile(vols, 75)) if len(vols) > 0 else 0
            iqrv = q3v - q1v
            lv, uv, nrv = q1v - 1.5 * iqrv, q3v + 1.5 * iqrv, 0

        # Clean TMO outliers
        if len(tmos) >= 4:
            tmo_clean, q1t, q3t, iqrt, lt, ut, nrt = remove_outliers_iqr(tmos)
        else:
            tmo_clean = tmos
            q1t = float(np.percentile(tmos, 25)) if len(tmos) > 0 else 0
            q3t = float(np.percentile(tmos, 75)) if len(tmos) > 0 else 0
            iqrt = q3t - q1t
            lt, ut, nrt = q1t - 1.5 * iqrt, q3t + 1.5 * iqrt, 0

        avg_vol = float(vol_clean.mean()) if len(vol_clean) > 0 else 0.0
        avg_tmo = float(tmo_clean.mean()) if len(tmo_clean) > 0 else 0.0

        pesos.append(avg_vol)
        fatores_tmo.append(avg_tmo)
        stats.append({
            "horario": h,
            "vol_media": round(avg_vol, 2),
            "vol_q1": round(q1v, 2), "vol_q3": round(q3v, 2),
            "vol_iqr": round(iqrv, 2),
            "vol_lower": round(lv, 2), "vol_upper": round(uv, 2),
            "vol_outliers": nrv,
            "vol_n": len(vols), "vol_n_clean": len(vol_clean),
            "tmo_media": round(avg_tmo, 2),
            "tmo_q1": round(q1t, 2), "tmo_q3": round(q3t, 2),
            "tmo_outliers": nrt,
            "tmo_n": len(tmos), "tmo_n_clean": len(tmo_clean),
        })

    # Normalize pesos to sum=100
    total_vol = sum(pesos)
    if total_vol > 0:
        pesos_pct = [p / total_vol * 100 for p in pesos]
    else:
        # Se zerou tudo (ex: sem funcionamento), distribui uniformemente dentro da janela ou zerado
        if window is not None and window[1] > window[0]:
            size = window[1] - window[0]
            pesos_pct = [0.0] * 48
            for i in range(window[0], window[1]):
                pesos_pct[i] = 100.0 / size
        else:
            pesos_pct = [0.0] * 48

    # Convert TMO to factors (relative to global average)
    global_tmo = sum(p * t for p, t in zip(pesos, fatores_tmo)) / (total_vol + 1e-9)
    if global_tmo > 0:
        tmo_factors = [t / global_tmo if t > 0 else 1.0 for t in fatores_tmo]
    else:
        tmo_factors = [1.0] * 48

    return pesos_pct, tmo_factors, stats, round(global_tmo, 2)


# ── Forecast: monthly volume ─────────────────────────────────────────
def forecast_volume_mensal(records: List[dict]):
    """
    Project monthly volume from historical data.
    Returns (projected_volume, stats).
    """
    # Aggregate by month
    by_month = {}
    for r in records:
        m = r["mes"]
        by_month[m] = by_month.get(m, 0) + r["volume"]

    if not by_month:
        return 0, {}

    months = sorted(by_month.keys())
    vols = np.array([by_month[m] for m in months], dtype=float)

    if len(vols) >= 4:
        clean, q1, q3, iqr, lb, ub, nr = remove_outliers_iqr(vols)
    else:
        clean = vols
        q1 = float(np.percentile(vols, 25)) if len(vols) > 1 else float(vols[0])
        q3 = float(np.percentile(vols, 75)) if len(vols) > 1 else float(vols[0])
        iqr = q3 - q1
        lb, ub, nr = q1 - 1.5 * iqr, q3 + 1.5 * iqr, 0

    avg = float(clean.mean()) if len(clean) > 0 else 0.0

    stats = {
        "meses": months,
        "volumes": [round(float(v)) for v in vols],
        "volume_projetado": round(avg),
        "q1": round(q1, 2), "q3": round(q3, 2),
        "iqr": round(iqr, 2),
        "outliers_removidos": nr,
        "n_meses": len(months),
    }
    return round(avg), stats


# ── Forecast: monthly TMO ────────────────────────────────────────────
def forecast_tmo_mensal(records: List[dict]):
    """
    Project monthly TMO (volume-weighted average after IQR cleaning).
    Returns (projected_tmo, stats).
    """
    # Per-day weighted TMO
    by_day = {}
    for r in records:
        d = r["data"]
        if d not in by_day:
            by_day[d] = {"vol": 0, "tmo_vol": 0}
        by_day[d]["vol"] += r["volume"]
        by_day[d]["tmo_vol"] += r["volume"] * r["tmo"]

    day_tmos = []
    for d, v in by_day.items():
        if v["vol"] > 0:
            day_tmos.append(v["tmo_vol"] / v["vol"])

    if not day_tmos:
        return 0, {}

    arr = np.array(day_tmos, dtype=float)
    if len(arr) >= 4:
        clean, q1, q3, iqr_val, lb, ub, nr = remove_outliers_iqr(arr)
    else:
        clean = arr
        q1 = float(np.percentile(arr, 25)) if len(arr) > 1 else float(arr[0])
        q3 = float(np.percentile(arr, 75)) if len(arr) > 1 else float(arr[0])
        iqr_val = q3 - q1
        nr = 0

    avg = float(clean.mean()) if len(clean) > 0 else 0.0

    stats = {
        "tmo_projetado": round(avg, 1),
        "q1": round(q1, 1), "q3": round(q3, 1),
        "outliers_removidos": nr,
        "n_dias": len(day_tmos),
    }
    return round(avg, 1), stats


# ── Forecast: daily curve ─────────────────────────────────────────────
def forecast_curva_diaria(records: List[dict], ano: int, mes: int, dias_ativos: List[str]):
    """
    Generate daily weight curve for target month using (weekday, ordinal) grouping.
    Returns list of DiaMes with peso_pct.
    """
    by_day = {}
    for r in records:
        d_str = r["data"]
        if d_str not in by_day:
            d = _dt.date.fromisoformat(d_str)
            ordinal = (d.day - 1) // 7 + 1
            if _is_feriado(d):
                bucket = "Feriado"
            else:
                bucket = (d.weekday(), ordinal)
            by_day[d_str] = {"vol": 0, "tipo": r["tipo_dia"], "bucket": bucket, "weekday": d.weekday()}
        by_day[d_str]["vol"] += r["volume"]

    # Group by bucket
    vols_by_bucket = {}
    vols_by_weekday = {i: [] for i in range(7)}
    for info in by_day.values():
        b = info["bucket"]
        v = info["vol"]
        if b not in vols_by_bucket:
            vols_by_bucket[b] = []
        vols_by_bucket[b].append(v)
        if b != "Feriado":
            vols_by_weekday[info["weekday"]].append(v)

    # Clean and average per bucket
    avg_bucket = {}
    for b, vols in vols_by_bucket.items():
        arr = np.array(vols, dtype=float)
        if len(arr) >= 4:
            clean, *_ = remove_outliers_iqr(arr)
        else:
            clean = arr
        avg_bucket[b] = float(clean.mean()) if len(clean) > 0 else 0.0

    # Also compute fallback averages for each weekday (in case a bucket is missing)
    avg_weekday = {}
    for wd, vols in vols_by_weekday.items():
        arr = np.array(vols, dtype=float)
        if len(arr) >= 4:
            clean, *_ = remove_outliers_iqr(arr)
        else:
            clean = arr
        avg_weekday[wd] = float(clean.mean()) if len(clean) > 0 else 0.0

    # Build target month calendar
    num_days = calendar.monthrange(ano, mes)[1]
    dias = []
    raw_weights = []

    # Map for dias_ativos
    wd_to_str = {0: "seg", 1: "ter", 2: "qua", 3: "qui", 4: "sex", 5: "sab", 6: "dom"}

    for d_num in range(1, num_days + 1):
        d = _dt.date(ano, mes, d_num)
        tipo = _tipo_dia(d)
        wd = d.weekday()
        
        # Check if day is active
        str_day = wd_to_str[wd]
        is_feriado_now = _is_feriado(d)
        
        if is_feriado_now and "fer" not in dias_ativos:
            w = 0.0
        elif not is_feriado_now and str_day not in dias_ativos:
            w = 0.0
        else:
            # Predict volume
            if is_feriado_now:
                w = avg_bucket.get("Feriado", avg_bucket.get((6, 1), 0.0)) # Fallback to Sunday if no holiday in history
            else:
                ordinal = (d.day - 1) // 7 + 1
                b = (wd, ordinal)
                if b in avg_bucket:
                    w = avg_bucket[b]
                else:
                    w = avg_weekday.get(wd, 0.0)
        
        raw_weights.append(w)
        dias.append({"data": f"{ano}-{mes:02d}-{d_num:02d}", "tipo": tipo})

    # Normalize to 100%
    total = sum(raw_weights)
    if total > 0:
        pesos = [w / total * 100 for w in raw_weights]
    else:
        # Se zerou tudo (ex: dias desativados), distribuir no pouco que tem?
        # Se total == 0, significa que nenhum dia ativo tem histórico. Deixa como 0.
        pesos = [0.0] * num_days

    result = []
    for i, dia in enumerate(dias):
        result.append(DiaMes(
            data=dia["data"],
            tipo=dia["tipo"],
            peso_pct=round(pesos[i], 4),
        ))

    return result


# ── Export forecast as Excel (same format as input template) ──────────
def generate_forecast_xlsx(
    curva_sem: Tuple[List[float], List[float]],
    curva_sab: Tuple[List[float], List[float]],
    curva_dom: Tuple[List[float], List[float]],
    dias: List[DiaMes],
    vol_mensal: int,
    tmo_mensal: float,
    stats_sem: List[dict],
    stats_sab: List[dict],
    stats_dom: List[dict],
) -> bytes:
    """
    Generate Excel in the same format as the input template
    (Curva_Semana, Curva_Sabado, Curva_Domingo, Curva_Dias)
    plus a Resumo sheet.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    h_fill = PatternFill("solid", fgColor="1a2030")
    h_font = Font(bold=True, color="4f9cf9", size=11)
    data_font = Font(color="cdd6f4", size=10)
    green_fill = PatternFill("solid", fgColor="071a0a")
    bd = Side(style="thin", color="1e2535")
    border = Border(left=bd, right=bd, top=bd, bottom=bd)
    center = Alignment(horizontal="center")

    def hdr(cell, text):
        cell.value = text
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center
        cell.border = border

    def dat(cell, value):
        cell.value = value
        cell.font = data_font
        cell.alignment = center
        cell.border = border

    # ── Intraday sheets ───────────────────────────────────────────────
    def make_intraday(ws, name, pesos, fatores):
        ws.title = name
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16

        hdr(ws.cell(row=2, column=1), "HORARIO")
        hdr(ws.cell(row=2, column=2), "PESO_PCT")
        hdr(ws.cell(row=2, column=3), "FATOR_TMO")

        for i, (h, p, f) in enumerate(zip(HORARIOS, pesos, fatores)):
            row = i + 3
            dat(ws.cell(row=row, column=1), h)
            dat(ws.cell(row=row, column=2), round(p, 4))
            dat(ws.cell(row=row, column=3), round(f, 4))
            if p >= 3.0:
                ws.cell(row=row, column=2).fill = green_fill
        ws.freeze_panes = "A3"

    ws1 = wb.active
    make_intraday(ws1, "Curva_Semana", curva_sem[0], curva_sem[1])
    make_intraday(wb.create_sheet(), "Curva_Sabado", curva_sab[0], curva_sab[1])
    make_intraday(wb.create_sheet(), "Curva_Domingo", curva_dom[0], curva_dom[1])

    # ── Dias sheet ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Curva_Dias")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 14

    hdr(ws4.cell(row=2, column=1), "DATA")
    hdr(ws4.cell(row=2, column=2), "TIPO")
    hdr(ws4.cell(row=2, column=3), "PESO_PCT")

    for i, d in enumerate(dias):
        row = i + 3
        dat(ws4.cell(row=row, column=1), d.data)
        dat(ws4.cell(row=row, column=2), d.tipo)
        dat(ws4.cell(row=row, column=3), round(d.peso_pct, 4))
    ws4.freeze_panes = "A3"

    # ── Resumo sheet ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Resumo_Forecast")
    ws5.sheet_view.showGridLines = False
    title_font = Font(bold=True, color="1f4e79", size=14)
    sub_font = Font(bold=True, color="000000", size=11)
    val_font = Font(color="000000", size=11)

    ws5["A1"] = "FORECAST — Resumo da Projeção"
    ws5["A1"].font = title_font

    ws5["A3"] = "Volume Mensal Projetado:"
    ws5["A3"].font = sub_font
    ws5["B3"] = vol_mensal
    ws5["B3"].font = val_font

    ws5["A4"] = "TMO Mensal Projetado (s):"
    ws5["A4"].font = sub_font
    ws5["B4"] = tmo_mensal
    ws5["B4"].font = val_font

    ws5["A6"] = "Este arquivo pode ser importado diretamente no WFM Engine como curvas de input."
    ws5["A6"].font = Font(color="6c7a96", size=10, italic=True)

    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Full forecast pipeline ────────────────────────────────────────────
def run_forecast(
    file_bytes: bytes, 
    ano_alvo: int, 
    mes_alvo: int,
    dias_ativos: Optional[List[str]] = None,
    window_util: Optional[Tuple[int, int]] = None,
    window_sab: Optional[Tuple[int, int]] = None,
    window_dom: Optional[Tuple[int, int]] = None,
) -> dict:
    """
    Full forecast pipeline:
    1. Parse historical data
    2. Clean outliers (IQR)
    3. Generate all curves
    4. Return results dict
    """
    if dias_ativos is None:
        dias_ativos = ["seg", "ter", "qua", "qui", "sex", "sab", "dom", "fer"]

    records = parse_historico_xlsx(file_bytes)
    if not records:
        raise ValueError("Nenhum dado válido encontrado no arquivo. "
                         "Verifique se a aba 'Historico' contém dados no formato correto.")

    # Intraday curves per day type
    pesos_sem, fat_sem, stats_sem, tmo_sem = forecast_curva_intraday(records, "Util", window=window_util)
    pesos_sab, fat_sab, stats_sab, tmo_sab = forecast_curva_intraday(records, "Sabado", window=window_sab)
    pesos_dom, fat_dom, stats_dom, tmo_dom = forecast_curva_intraday(records, "Domingo", window=window_dom)

    # Monthly projections
    vol_mensal, vol_stats = forecast_volume_mensal(records)
    tmo_mensal, tmo_stats = forecast_tmo_mensal(records)

    # Daily curve
    dias = forecast_curva_diaria(records, ano_alvo, mes_alvo, dias_ativos)

    # Total outliers removed
    total_outliers = (
        sum(s["vol_outliers"] + s["tmo_outliers"] for s in stats_sem) +
        sum(s["vol_outliers"] + s["tmo_outliers"] for s in stats_sab) +
        sum(s["vol_outliers"] + s["tmo_outliers"] for s in stats_dom) +
        vol_stats.get("outliers_removidos", 0) +
        tmo_stats.get("outliers_removidos", 0)
    )

    # Count records by type
    n_util = sum(1 for r in records if r["tipo_dia"] == "Util")
    n_sab = sum(1 for r in records if r["tipo_dia"] == "Sabado")
    n_dom = sum(1 for r in records if r["tipo_dia"] == "Domingo")

    return {
        "curva_semana": {"pesos": pesos_sem, "fatores_tmo": fat_sem, "tmo_base": tmo_sem},
        "curva_sabado": {"pesos": pesos_sab, "fatores_tmo": fat_sab, "tmo_base": tmo_sab},
        "curva_domingo": {"pesos": pesos_dom, "fatores_tmo": fat_dom, "tmo_base": tmo_dom},
        "dias": [{"data": d.data, "tipo": d.tipo, "peso_pct": d.peso_pct} for d in dias],
        "volume_mensal": vol_mensal,
        "tmo_mensal": tmo_mensal,
        "stats_semana": stats_sem,
        "stats_sabado": stats_sab,
        "stats_domingo": stats_dom,
        "vol_stats": vol_stats,
        "tmo_stats": tmo_stats,
        "total_outliers": total_outliers,
        "n_registros": len(records),
        "n_util": n_util,
        "n_sabado": n_sab,
        "n_domingo": n_dom,
        "ano_alvo": ano_alvo,
        "mes_alvo": mes_alvo,
        # Keep raw for export
        "_curvas": {
            "sem": (pesos_sem, fat_sem),
            "sab": (pesos_sab, fat_sab),
            "dom": (pesos_dom, fat_dom),
        },
        "_dias": dias,
        "_stats": {
            "sem": stats_sem, "sab": stats_sab, "dom": stats_dom,
        },
    }
