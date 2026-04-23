"""
Excel export and validation for WFM Engine.
"""
import io
from typing import Tuple, List
from .models import WFMOutput

# ── Validation ────────────────────────────────────────────────────────────
def validate_excel(file_bytes: bytes) -> Tuple[bool, List[dict]]:
    """
    Validate uploaded Excel.
    Returns (is_valid, list of issues).
    Each issue: {sheet, row, col, severity, message}
    """
    import openpyxl
    issues = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        return False, [{"sheet":"—","row":0,"col":"—","severity":"erro",
                        "message":f"Arquivo inválido ou corrompido: {e}"}]

    expected_sheets = ["Curva_Semana","Curva_Sabado","Curva_Domingo"]
    for sheet in expected_sheets:
        if sheet not in wb.sheetnames:
            issues.append({"sheet":sheet,"row":0,"col":"—","severity":"aviso",
                           "message":f"Aba '{sheet}' não encontrada — será usada curva padrão"})
            continue
        ws = wb[sheet]
        rows_data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None: continue
            rows_data.append(row)

        if len(rows_data) != 48:
            issues.append({"sheet":sheet,"row":0,"col":"—","severity":"erro",
                           "message":f"Deve ter 48 linhas de dados, encontrou {len(rows_data)}"})
            continue

        pesos = []
        for i, row in enumerate(rows_data, 3):
            # peso_pct
            try:
                p = float(str(row[1]).replace(',','.')) if row[1] is not None else None
                if p is None:
                    issues.append({"sheet":sheet,"row":i,"col":"PESO_PCT","severity":"erro",
                                   "message":"Célula vazia"})
                elif p < 0:
                    issues.append({"sheet":sheet,"row":i,"col":"PESO_PCT","severity":"erro",
                                   "message":f"Valor negativo: {p}"})
                else:
                    pesos.append(p)
            except ValueError:
                issues.append({"sheet":sheet,"row":i,"col":"PESO_PCT","severity":"erro",
                               "message":f"Valor não numérico: '{row[1]}'"})
            # fator_tmo
            try:
                f = float(str(row[2]).replace(',','.')) if row[2] is not None else 1.0
                if f < 0.3 or f > 5.0:
                    issues.append({"sheet":sheet,"row":i,"col":"FATOR_TMO","severity":"aviso",
                                   "message":f"Valor fora do range esperado (0.3–5.0): {f}"})
            except ValueError:
                issues.append({"sheet":sheet,"row":i,"col":"FATOR_TMO","severity":"aviso",
                               "message":f"Valor não numérico: '{row[2]}' — usando 1.0"})

        if pesos:
            soma = sum(pesos)
            if soma < 98 or soma > 102:
                issues.append({"sheet":sheet,"row":0,"col":"PESO_PCT","severity":"erro",
                               "message":f"Soma dos pesos = {soma:.2f}% (deve ser 100%)"})
            elif soma < 99 or soma > 101:
                issues.append({"sheet":sheet,"row":0,"col":"PESO_PCT","severity":"aviso",
                               "message":f"Soma dos pesos = {soma:.2f}% (idealmente exatamente 100%)"})

    # Curva_Dias (optional)
    if "Curva_Dias" in wb.sheetnames:
        ws = wb["Curva_Dias"]
        tipos_validos = {"Util","Sabado","Domingo"}
        pesos_dias = []
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            if row[0] is None: continue
            if str(row[1]).strip() not in tipos_validos:
                issues.append({"sheet":"Curva_Dias","row":i,"col":"TIPO","severity":"erro",
                               "message":f"Tipo inválido: '{row[1]}' — use Util, Sabado ou Domingo"})
            try:
                p = float(str(row[2]).replace(',','.')) if row[2] is not None else None
                if p is not None: pesos_dias.append(p)
            except ValueError:
                issues.append({"sheet":"Curva_Dias","row":i,"col":"PESO_PCT","severity":"erro",
                               "message":f"Valor não numérico: '{row[2]}'"})
        if pesos_dias:
            soma = sum(pesos_dias)
            if soma < 98 or soma > 102:
                issues.append({"sheet":"Curva_Dias","row":0,"col":"PESO_PCT","severity":"erro",
                               "message":f"Soma dos pesos dos dias = {soma:.2f}% (deve ser 100%)"})

    has_error = any(i["severity"]=="erro" for i in issues)
    return not has_error, issues


# ── Export ────────────────────────────────────────────────────────────────
def export_resultado_xlsx(out: WFMOutput) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Styles — paleta clara com alto contraste p/ leitura em tela e impressão
    h_fill  = PatternFill("solid", fgColor="1f4e79")
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    ok_col  = "006100"
    bad_col = "9C0006"
    dat_col = "000000"
    tot_fill= PatternFill("solid", fgColor="FFF2CC")
    tot_font= Font(bold=True, color="000000", size=11)
    dat_font= Font(color=dat_col, size=11)
    alt_fill= PatternFill("solid", fgColor="F2F2F2")
    bd = Side(style="thin", color="808080")
    bdr = Border(left=bd, right=bd, top=bd, bottom=bd)
    ctr = Alignment(horizontal="center", vertical="center")

    def hdr(cell, text):
        cell.value=text; cell.font=h_font; cell.fill=h_fill
        cell.alignment=ctr; cell.border=bdr

    def dat(cell, value, color=None):
        cell.value=value
        cell.font=Font(color=color or dat_col, size=11)
        cell.alignment=ctr; cell.border=bdr

    # ── Sheet 1: Resumo Mensal ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo Mensal"
    ws1.sheet_view.showGridLines = False

    # Header info
    ws1['A1'] = f"WFM Engine — Dimensionamento {out.mes:02d}/{out.ano}"
    ws1['A1'].font = Font(bold=True, color="1f4e79", size=14)
    ws1['A2'] = f"SLA Mensal: {out.sla_ponderado_mes*100:.1f}%  |  HC Bruto Total: {out.hc_fisico['total']}  |  Volume: {out.volume_total_mes:,.0f}  |  NS: {out.ns_total_mes:,.0f}"
    ws1['A2'].font = Font(bold=True, color="000000", size=11)

    # Pool summary
    ws1['A4'] = "POOL"; ws1['B4'] = "TIPO"; ws1['C4'] = "HC BRUTO"; ws1['D4'] = "HC LÍQUIDO REF"
    ws1['E4'] = "PL BASE"; ws1['F4'] = "PL EFETIVO"
    for col in "ABCDEF":
        hdr(ws1[f'{col}4'], ws1[f'{col}4'].value)
    rows_pool = [
        ("Pool 8:12","8:12",out.hc_fisico.get("pool_812",0),out.hc_liquido_ref.get("pool_812",0),out.pl_efetivo.get("8:12_base","—"),out.pl_efetivo.get("8:12","—")),
        ("Pool B-Sáb","6:20",out.hc_fisico.get("pool_b_sab",0),out.hc_liquido_ref.get("pool_b_sab",0),out.pl_efetivo.get("6:20_base","—"),out.pl_efetivo.get("6:20","—")),
        ("Pool B-Dom","6:20",out.hc_fisico.get("pool_b_dom",0),out.hc_liquido_ref.get("pool_b_dom",0),out.pl_efetivo.get("6:20_base","—"),out.pl_efetivo.get("6:20","—")),
        ("TOTAL","—",out.hc_fisico.get("total",0),out.hc_liquido_ref.get("total",0),"—","—"),
    ]
    for i, r in enumerate(rows_pool, 5):
        for j, v in enumerate(r, 1):
            c = ws1.cell(row=i, column=j, value=v)
            c.font = tot_font if r[0]=="TOTAL" else dat_font
            if r[0]=="TOTAL":
                c.fill = tot_fill
            c.alignment = ctr; c.border = bdr

    # Day summary table
    row_start = 11
    cols_day = ["DATA","DIA","TIPO","VOLUME","TMO(s)","HC LÍQ","HC BRUTO","SLA%","NS","TME(s)","OCUPAÇÃO%","FILA%","ABANDONO%","INTERVALOS OK%","STATUS"]
    for j, col in enumerate(cols_day, 1):
        hdr(ws1.cell(row=row_start, column=j), col)

    for i, d in enumerate(out.dias, row_start+1):
        # Compute volume-weighted abandono for the day
        vol_total_day = sum(iv.volume for iv in d.intervalos) if d.intervalos else 0
        abd_day = (sum(iv.volume * iv.p_abandon * 100 for iv in d.intervalos) / vol_total_day) if vol_total_day > 0 else 0.0
        vals = [d.data, d.dia_semana, d.tipo, d.volume_total, d.tmo_medio,
                d.hc_liq_max, d.hc_bruto_max,
                round(d.sla_ponderado*100,1), d.ns_total, d.tme_medio,
                round(d.ocupacao_media*100,1), round(d.fila_media*100,1),
                round(abd_day,2),
                round(d.intervalos_ok_pct*100,1),
                "OK" if d.status_sla=="ok" else "Abaixo"]
        zebra = (i % 2 == 0)
        for j, v in enumerate(vals, 1):
            c = ws1.cell(row=i, column=j, value=v)
            c.alignment = ctr; c.border = bdr
            if zebra: c.fill = alt_fill
            if j == 15:
                c.font = Font(bold=True, color=ok_col if d.status_sla=="ok" else bad_col, size=11)
            elif j == 8:
                c.font = Font(bold=True, color=ok_col if d.sla_ponderado >= out.sla_target else bad_col, size=11)
            else:
                c.font = dat_font

    # Total row
    tot_row = row_start + len(out.dias) + 1
    totals = ["TOTAL MÊS","—","—",
              out.volume_total_mes, "—", "—", "—",
              round(out.sla_ponderado_mes*100,1), out.ns_total_mes, "—","—","—","—",
              round((out.intervalos_ok_pct or 0)*100,1) if out.intervalos_ok_pct else "—",
              "OK" if out.sla_ponderado_mes >= out.sla_target else "Abaixo"]
    for j, v in enumerate(totals, 1):
        c = ws1.cell(row=tot_row, column=j, value=v)
        c.font = tot_font; c.fill = tot_fill; c.alignment = ctr; c.border = bdr

    # Auto width
    for col in ws1.columns:
        max_w = max((len(str(cell.value or "")) for cell in col), default=8)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 30)

    # ── Sheet 2: Turnos ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Turnos")
    ws2.sheet_view.showGridLines = False
    for j, h in enumerate(["POOL","TIPO","ENTRADA","SAÍDA","HC BRUTO","HC LÍQUIDO REF"], 1):
        hdr(ws2.cell(row=1, column=j), h)
    for i, t in enumerate(out.turnos, 2):
        zebra = (i % 2 == 0)
        for j, v in enumerate([t.pool,t.tipo,t.entrada,t.saida,t.agentes_bruto,t.agentes_liq], 1):
            c = ws2.cell(row=i, column=j, value=v); c.font=dat_font; c.alignment=ctr; c.border=bdr
            if zebra: c.fill = alt_fill
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 18

    # ── Sheet 3: Detalhe por Intervalo (todos os dias em uma única aba) ──
    ws = wb.create_sheet("Detalhe Intervalos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    cols_iv = ["DATA","DIA","TIPO","HORÁRIO","VOLUME","TMO(s)","HC LÍQUIDO","HC BRUTO",
               "TRÁFEGO(Erl)","FILA(Pw%)","TME(s)","OCUPAÇÃO%","SLA%","NS","ABANDONO%","STATUS"]
    for j, h in enumerate(cols_iv, 1):
        hdr(ws.cell(row=1, column=j), h)
    row = 2
    for di, dia in enumerate(out.dias):
        if not dia.intervalos: continue
        day_fill = PatternFill("solid", fgColor="FFF2CC" if di % 2 == 0 else "FCE4D6")
        # Cabeçalho do dia
        c = ws.cell(row=row, column=1,
                    value=f"{dia.data} {dia.dia_semana} · {dia.tipo} · Vol {dia.volume_total:,.0f} · SLA {dia.sla_ponderado*100:.1f}% · NS {dia.ns_total:,.0f}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols_iv))
        c.font = Font(bold=True, color="1f4e79", size=12)
        c.fill = day_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1
        for iv in dia.intervalos:
            ok = iv.sla_pct >= out.sla_target
            vals = [dia.data, dia.dia_semana, dia.tipo,
                    iv.horario, iv.volume, iv.tmo, iv.hc_liq, iv.hc_bruto,
                    iv.trafico_erl, round(iv.fila_pw*100,2), iv.tme_seg,
                    round(iv.ocupacao*100,2), round(iv.sla_pct*100,2), iv.ns,
                    round(iv.p_abandon*100,2),
                    "OK" if ok else "Abaixo"]
            zebra = (row % 2 == 0)
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=j, value=v)
                if j in (13, 16):
                    c.font = Font(bold=True, color=ok_col if ok else bad_col, size=11)
                else:
                    c.font = dat_font
                c.alignment = ctr; c.border = bdr
                if zebra: c.fill = alt_fill
            row += 1
    widths = [12,8,10,10,10,9,12,11,13,11,10,13,10,10,11,12]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
