"""
Demand distribution + Excel template generation + file parsing.
"""
import math, io, calendar
from typing import List
import numpy as np
from .models import CurvaIntraday, DiaMes, INTERVALS_PER_DAY, INTERVALS_30MIN

# ── Horários de 30min ─────────────────────────────────────────────────────
HORARIOS = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]

# ── Default curves ────────────────────────────────────────────────────────
def _bimodal(n=48):
    raw = [math.exp(-0.5*((i/2-10)/1.5)**2) + 0.8*math.exp(-0.5*((i/2-15)/1.8)**2) for i in range(n)]
    s = sum(raw); return [max(v/s*100, 0.001) for v in raw]

def _morning(n=48):
    raw = [math.exp(-0.5*((i/2-9)/2)**2) for i in range(n)]
    s = sum(raw); return [max(v/s*100, 0.001) for v in raw]

def _flat(n=48):
    raw = [1.0 if 8 <= i/2 < 18 else 0.001 for i in range(n)]
    s = sum(raw); return [max(v/s*100, 0.001) for v in raw]

def make_bimodal() -> List[float]: return _bimodal()
def make_morning() -> List[float]: return _morning()
def make_flat_business() -> List[float]: return _flat()

def default_curva(tipo: str) -> CurvaIntraday:
    pesos = {"Util": _bimodal(), "Sabado": _morning(), "Domingo": _flat()}.get(tipo, _bimodal())
    return CurvaIntraday(pesos=pesos, fatores_tmo=[1.0]*48)

# ── Normalise ─────────────────────────────────────────────────────────────
def normalise(pesos: List[float]) -> np.ndarray:
    arr = np.array(pesos, dtype=float)
    s = arr.sum()
    if s <= 0: arr[:] = 1.0; s = float(len(arr))
    return arr / s

# ── Expand 30min → 10min ─────────────────────────────────────────────────
def expand_to_10min(curve_30: np.ndarray) -> np.ndarray:
    out = np.zeros(INTERVALS_PER_DAY)
    for i in range(INTERVALS_30MIN):
        out[i*3:i*3+3] = curve_30[i] / 3
    return out

# ── Volume + TMO curves for a day ────────────────────────────────────────
def build_day_curves(vol_dia: float, curva: CurvaIntraday, tmo_base: float):
    pesos_norm = normalise(curva.pesos)
    vol_30m = pesos_norm * vol_dia
    vol_10m = expand_to_10min(vol_30m)
    tmo_10m = np.zeros(INTERVALS_PER_DAY)
    for i in range(INTERVALS_30MIN):
        tmo_10m[i*3:i*3+3] = tmo_base * curva.fatores_tmo[i]
    return vol_10m, tmo_10m

# ── Daily volumes ─────────────────────────────────────────────────────────
def volume_from_peso(peso_pct: float, volume_mes: float) -> float:
    return volume_mes * peso_pct / 100.0

# ── Auto-detect operating hours from curve ────────────────────────────────
def detect_operating_hours(curva: CurvaIntraday, threshold_pct: float = 0.1):
    """
    Returns (first_slot, last_slot) in 10-min units.
    If volume exists in early morning (before 05:00), treats as 24h operation.
    """
    pesos = curva.pesos  # 48 values summing to 100
    # Check if 24h: any volume before 05:00 (slot 10) or after 22:00 (slot 44)
    early_vol = sum(pesos[:10])  # 00:00-04:30
    late_vol  = sum(pesos[44:])  # 22:00-23:30
    if early_vol > 0.5 or late_vol > 0.5:
        return 0, 144  # full 24h
    # Otherwise detect by threshold
    first_30 = next((i for i, p in enumerate(pesos) if p >= threshold_pct), 0)
    last_30  = next((i for i, p in reversed(list(enumerate(pesos))) if p >= threshold_pct), 47)
    return first_30 * 3, (last_30 + 1) * 3

# ── Auto-generate dias_mes ────────────────────────────────────────────────
def gen_dias_mes(ano, mes, dias_uteis, vol_sab_pct, vol_dom_pct, volume_mes) -> List[DiaMes]:
    ano = int(ano); mes = int(mes)
    if not (1 <= mes <= 12):
        raise ValueError(f"Mês inválido: {mes}")
    num_days = calendar.monthrange(ano, mes)[1]
    tipos = []
    for d in range(1, num_days+1):
        dow = calendar.weekday(ano, mes, d)
        tipos.append("Sabado" if dow == 5 else "Domingo" if dow == 6 else "Util")
    util_count = tipos.count("Util")
    sab_count  = tipos.count("Sabado")
    dom_count  = tipos.count("Domingo")
    denom = util_count + sab_count*vol_sab_pct + dom_count*vol_dom_pct
    peso_util = 100.0 / denom if denom > 0 else 0
    dias = []
    for d, tipo in enumerate(tipos, 1):
        peso = peso_util if tipo=="Util" else peso_util*vol_sab_pct if tipo=="Sabado" else peso_util*vol_dom_pct
        dias.append(DiaMes(data=f"{ano}-{mes:02d}-{d:02d}", tipo=tipo, peso_pct=round(peso, 4)))
    return dias

# ── Excel template generation ─────────────────────────────────────────────
def get_templates_xlsx() -> bytes:
    """
    Single Excel file with 4 sheets:
      - Curva_Semana
      - Curva_Sabado
      - Curva_Domingo
      - Curva_Dias
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Style helpers ────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1a2030")
    header_font = Font(bold=True, color="4f9cf9", size=11)
    inst_font   = Font(color="6c7a96", size=10, italic=True)
    data_font   = Font(color="cdd6f4", size=10)
    green_fill  = PatternFill("solid", fgColor="071a0a")
    border_side = Side(style="thin", color="1e2535")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center      = Alignment(horizontal="center")

    def style_header(cell, text):
        cell.value = text
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = center
        cell.border = thin_border

    def style_data(cell, value):
        cell.value = value
        cell.font  = data_font
        cell.alignment = center
        cell.border = thin_border

    # ── Intraday sheet ───────────────────────────────────────────────────
    def make_intraday_sheet(ws, nome, pesos_default):
        ws.title = nome
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 45

        # Instructions block
        inst_rows = [
            ("INSTRUÇÕES DE PREENCHIMENTO",),
            ("",),
            ("Coluna HORARIO: não altere — são os 48 horários do dia (00:00 a 23:30).",),
            ("Coluna PESO_PCT: % do volume diário que chega nesse horário.",),
            ("  → A soma de todos os 48 valores DEVE ser 100.",),
            ("  → Use vírgula como decimal. Ex: 8,5 significa 8.5%",),
            ("  → Horários sem movimento: coloque 0",),
            ("Coluna FATOR_TMO: quanto o TMO varia nesse horário.",),
            ("  → 1,0 = TMO normal",),
            ("  → 1,1 = TMO 10% maior que o base",),
            ("  → 0,9 = TMO 10% menor que o base",),
            ("  → Se não souber, deixe 1,0 em todos",),
        ]
        for r, row in enumerate(inst_rows, 1):
            cell = ws.cell(row=r, column=4, value=row[0])
            cell.font = inst_font
            if r == 1:
                cell.font = Font(color="f9e2af", size=11, bold=True)

        # Headers
        style_header(ws.cell(row=2, column=1), "HORARIO")
        style_header(ws.cell(row=2, column=2), "PESO_PCT")
        style_header(ws.cell(row=2, column=3), "FATOR_TMO")

        # Data rows
        for i, (h, p) in enumerate(zip(HORARIOS, pesos_default)):
            row = i + 3
            style_data(ws.cell(row=row, column=1), h)
            style_data(ws.cell(row=row, column=2), round(p, 4))
            style_data(ws.cell(row=row, column=3), 1.0)
            # Highlight if peak
            if p >= 3.0:
                ws.cell(row=row, column=2).fill = green_fill

        # Freeze header
        ws.freeze_panes = "A3"

    # ── Dias sheet ───────────────────────────────────────────────────────
    def make_dias_sheet(ws):
        ws.title = "Curva_Dias"
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 50

        inst_rows = [
            ("INSTRUÇÕES DE PREENCHIMENTO",),
            ("",),
            ("Coluna DATA: data no formato AAAA-MM-DD. Ex: 2025-05-01",),
            ("Coluna TIPO: escreva exatamente Util, Sabado ou Domingo",),
            ("Coluna PESO_PCT: % do volume MENSAL que chega nesse dia.",),
            ("  → A soma de TODOS os dias DEVE ser 100.",),
            ("  → Dias úteis costumam ter peso entre 3% e 6%",),
            ("  → Sábados entre 1% e 3%, Domingos entre 0.5% e 2%",),
            ("  → Se não tiver essa informação, deixe a aba em branco",),
            ("    e o sistema vai calcular automaticamente.",),
        ]
        for r, row in enumerate(inst_rows, 1):
            cell = ws.cell(row=r, column=4, value=row[0])
            cell.font = inst_font
            if r == 1:
                cell.font = Font(color="f9e2af", size=11, bold=True)

        style_header(ws.cell(row=2, column=1), "DATA")
        style_header(ws.cell(row=2, column=2), "TIPO")
        style_header(ws.cell(row=2, column=3), "PESO_PCT")

        # Example month (May 2025)
        dias = gen_dias_mes(2025, 5, 22, 0.35, 0.15, 100)
        for i, d in enumerate(dias):
            row = i + 3
            style_data(ws.cell(row=row, column=1), d.data)
            style_data(ws.cell(row=row, column=2), d.tipo)
            style_data(ws.cell(row=row, column=3), round(d.peso_pct, 4))

        ws.freeze_panes = "A3"

    # ── Build sheets ─────────────────────────────────────────────────────
    ws1 = wb.active
    make_intraday_sheet(ws1, "Curva_Semana",  _bimodal())
    make_intraday_sheet(wb.create_sheet(), "Curva_Sabado",  _morning())
    make_intraday_sheet(wb.create_sheet(), "Curva_Domingo", _flat())
    make_dias_sheet(wb.create_sheet())

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── Parse uploaded Excel ──────────────────────────────────────────────────
def parse_curva_xlsx(wb, sheet_name: str) -> CurvaIntraday:
    import openpyxl
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada no arquivo.")
    ws = wb[sheet_name]
    pesos, fatores = [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None: continue
        try:
            p = float(str(row[1]).replace(',', '.')) if row[1] is not None else 0.0
            f = float(str(row[2]).replace(',', '.')) if row[2] is not None else 1.0
            pesos.append(p)
            fatores.append(f)
        except (ValueError, TypeError):
            continue
    if len(pesos) != 48:
        raise ValueError(f"Aba '{sheet_name}' deve ter 48 linhas de dados, encontrou {len(pesos)}.")
    s = sum(pesos) or 1
    pesos = [p/s*100 for p in pesos]
    return CurvaIntraday(pesos=pesos, fatores_tmo=fatores)

def parse_dias_xlsx(wb) -> List[DiaMes]:
    """Read dias. TIPO derived from calendar, ignores what user typed."""
    if "Curva_Dias" not in wb.sheetnames:
        return []
    import datetime as _dt
    ws = wb["Curva_Dias"]
    dias = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None: continue
        try:
            data_str = str(row[0]).strip()
            d = _dt.date.fromisoformat(data_str)
            dow = d.weekday()
            tipo = "Sabado" if dow == 5 else "Domingo" if dow == 6 else "Util"
            peso = float(str(row[2]).replace(",", ".")) if row[2] is not None else 0.0
            dias.append(DiaMes(data=data_str, tipo=tipo, peso_pct=peso))
        except (ValueError, TypeError):
            continue
    return dias


def extract_mes_ano_from_xlsx(wb):
    """Extract mes/ano from first date in Curva_Dias. Returns (mes, ano) or None."""
    if "Curva_Dias" not in wb.sheetnames:
        return None
    import datetime as _dt
    ws = wb["Curva_Dias"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None: continue
        try:
            d = _dt.date.fromisoformat(str(row[0]).strip())
            return d.month, d.year
        except (ValueError, TypeError):
            continue
    return None

def parse_excel_upload(file_bytes: bytes):
    """Parse all sheets from uploaded Excel. Returns (curva_sem, curva_sab, curva_dom, dias)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    curva_sem = parse_curva_xlsx(wb, "Curva_Semana")  if "Curva_Semana"  in wb.sheetnames else None
    curva_sab = parse_curva_xlsx(wb, "Curva_Sabado")  if "Curva_Sabado"  in wb.sheetnames else None
    curva_dom = parse_curva_xlsx(wb, "Curva_Domingo") if "Curva_Domingo" in wb.sheetnames else None
    dias      = parse_dias_xlsx(wb)
    return curva_sem, curva_sab, curva_dom, dias

# ── Slot helpers ──────────────────────────────────────────────────────────
def slot_to_time(slot: int) -> str:
    m = slot * 10
    return f"{m//60:02d}:{m%60:02d}"

def time_to_slot(time_str: str) -> int:
    """Convert 'HH:MM' to 10-min slot index."""
    h, m = map(int, time_str.split(':'))
    return (h * 60 + m) // 10

def to_30min_mean(arr: np.ndarray) -> List[float]:
    return [float(arr[i*3:i*3+3].mean()) for i in range(INTERVALS_30MIN)]
